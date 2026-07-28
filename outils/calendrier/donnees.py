# -*- coding: utf-8 -*-
# ⚠️ DEPOT : VeVePreda/scrapeur-veve   ·   CHEMIN : outils/calendrier/donnees.py
# Le projet vit sur 6 depots et DEUX comptes GitHub. Un fichier
# depose au mauvais endroit ne provoque aucune erreur : il dort.

"""📅 LES DROPS D'UNE FENETRE, A LA MAILLE **SERIE** (jamais a la maille rarete).

D'OU VIENT LA DONNEE
--------------------
Des onglets 🟢C-COMICS / 🔵C-COLLECTIBLE — exactement la meme source que
`scraper/discord_drops.py`. **Aucune requete vers VeVe** : le calendrier ne peut
pas dire autre chose que le Sheet, et il ne se fait remarquer d'aucune source.

Deux lectures possibles :
  1. `--sheet`  : le Google Sheet en direct (via `scraper.sheets`, secrets requis) ;
  2. `--xlsx X` : le miroir `VeVe Scraper.xlsx` — **hors reseau**, mode par
     defaut en local et le seul qui marche dans un bac a sable.

⚠️ LA MAILLE EST LA SERIE, PAS LA LIGNE
---------------------------------------
Un comic = 5 lignes (COMMON → SECRET_RARE) qui sont **UN SEUL drop**. Compter
les lignes ferait afficher « 150 drops » un mercredi ou il y en a 30. On
regroupe donc par `series_uuid` (repli : nom de serie, puis nom d'element) et on
choisit UNE image representative — celle de la rarete la plus BASSE, la
couverture de base, celle que tout le monde reconnait. Le choix est
**deterministe** : deux runs le meme jour donnent la meme image.

⚠️ CE QUE LA FENETRE MONTRE VRAIMENT
------------------------------------
VeVe n'annonce ses drops que ~5 jours a l'avance. Une fenetre tournee vers
l'AVENIR serait vide a 90 %. La fenetre par defaut est donc celle du visuel de
reference : **5 semaines qui SE TERMINENT par la semaine en cours** (4 semaines
passees + la semaine courante), du lundi au dimanche. C'est un « ou en est-on »,
et c'est ce qui est poste le samedi.
`--decalage N` deplace la fenetre de N semaines (negatif = vers le passe).
"""

from __future__ import annotations

import datetime as _dt
import os
from dataclasses import dataclass, field
from typing import Dict, Iterable, List, Optional

# Onglets sources : (nom d'onglet, nature)
ONGLETS = (("🟢C-COMICS", "comic"), ("🔵C-COLLECTIBLE", "collectible"))

# Ordre des raretes, du plus commun au plus rare. Sert a choisir la couverture
# representative d'une serie. Une valeur inconnue tombe a la fin (999).
ORDRE_RARETE = {
    "COMMON": 0, "UNCOMMON": 1, "RARE": 2, "ULTRA_RARE": 3,
    "SECRET_RARE": 4, "ARTIST_PROOF": 5, "EXCLUSIVE": 6,
}

# Colonnes lues. Toute colonne absente d'un onglet vaut "" — les deux onglets
# n'ont PAS les memes colonnes.
COLONNES = ("veve_uuid", "name", "rarity", "edition_type", "releaseDate",
            "veve_series_name", "series_uuid", "veve_brand", "veve_licensor",
            "image_url", "drop_method", "supply")


# --------------------------------------------------------------------- modeles

@dataclass
class Serie:
    """Un drop, a la maille serie : ce qui merite UNE vignette."""
    cle: str                      # series_uuid, ou repli
    nom: str                      # nom affichable (serie si connue, sinon element)
    marque: str
    licensor: str
    nature: str                   # "comic" | "collectible"
    image_url: str
    drop_method: str
    editions: int = 0             # nb de lignes (raretes) regroupees
    rang_image: int = 999         # rarete de la ligne qui a fourni l'image


@dataclass
class Jour:
    date: _dt.date
    series: List[Serie] = field(default_factory=list)

    @property
    def vide(self) -> bool:
        return not self.series

    @property
    def nb(self) -> int:
        return len(self.series)


# --------------------------------------------------------------------- fenetre

def lundi(d: _dt.date) -> _dt.date:
    """Le lundi de la semaine de `d` (ISO : lundi = 0)."""
    return d - _dt.timedelta(days=d.weekday())


def fenetre(aujourdhui: _dt.date, semaines: int = 5, decalage: int = 0):
    """(debut, fin) — `semaines` semaines qui SE TERMINENT par celle d'aujourd'hui.

    `decalage` deplace le tout de N semaines (negatif = vers le passe).
    Retourne un lundi et un dimanche, bornes incluses.
    """
    if semaines < 1:
        raise ValueError("il faut au moins une semaine")
    derniere = lundi(aujourdhui) + _dt.timedelta(weeks=decalage)
    debut = derniere - _dt.timedelta(weeks=semaines - 1)
    return debut, derniere + _dt.timedelta(days=6)


def jours_de(debut: _dt.date, fin: _dt.date) -> List[_dt.date]:
    return [debut + _dt.timedelta(days=i) for i in range((fin - debut).days + 1)]


# ---------------------------------------------------------------------- lecture

def date_de(valeur) -> Optional[_dt.date]:
    """La `releaseDate` telle qu'elle sort du Sheet ou du xlsx.

    ⚠️ Le Sheet la rend en TEXTE (`2026-07-22 20:00:00`), le xlsx en datetime.
    Les deux passent ici, et rien d'autre : une date illisible fait sauter la
    ligne, jamais planter le run.
    """
    if valeur is None or valeur == "":
        return None
    if isinstance(valeur, _dt.datetime):
        return valeur.date()
    if isinstance(valeur, _dt.date):
        return valeur
    txt = str(valeur).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
                "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return _dt.datetime.strptime(txt[:len(fmt) + 2].strip(), fmt).date()
        except ValueError:
            continue
    try:                                    # ISO avec fuseau (…+00:00, …Z)
        return _dt.datetime.fromisoformat(txt.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def lignes_xlsx(chemin: str) -> List[dict]:
    """Les lignes des deux onglets depuis le miroir `VeVe Scraper.xlsx`."""
    import openpyxl                                     # import tardif : optionnel
    if not chemin or not os.path.exists(chemin):
        raise FileNotFoundError(
            "miroir introuvable : %r\n"
            "→ passer --xlsx <chemin de 'VeVe Scraper.xlsx'>, ou --sheet pour "
            "lire le Google Sheet en direct." % chemin)
    classeur = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    sorties: List[dict] = []
    for onglet, nature in ONGLETS:
        if onglet not in classeur.sheetnames:
            continue
        feuille = classeur[onglet]
        iterateur = feuille.iter_rows(values_only=True)
        entete = [str(c) if c is not None else "" for c in next(iterateur)]
        pos = {nom: i for i, nom in enumerate(entete)}
        for ligne in iterateur:
            if not ligne or not ligne[0]:
                continue
            item = {c: (ligne[pos[c]] if c in pos and pos[c] < len(ligne) else "")
                    for c in COLONNES}
            item["_nature"] = nature
            sorties.append(item)
    classeur.close()
    return sorties


def lignes_sheet() -> List[dict]:
    """Les memes lignes, mais depuis le Google Sheet en direct."""
    from scraper.sheets import _client                  # import tardif : secrets
    classeur = _client()
    sorties: List[dict] = []
    for onglet, nature in ONGLETS:
        feuille = classeur.worksheet(onglet)
        for enreg in feuille.get_all_records():
            item = {c: enreg.get(c, "") for c in COLONNES}
            item["_nature"] = nature
            sorties.append(item)
    return sorties


# -------------------------------------------------------------------- regroupage

def cle_serie(ligne: dict) -> str:
    for champ in ("series_uuid", "veve_series_name", "name"):
        valeur = str(ligne.get(champ) or "").strip()
        if valeur:
            return valeur
    return ""


def grouper(lignes: Iterable[dict], debut: _dt.date, fin: _dt.date) -> Dict[_dt.date, Jour]:
    """{date -> Jour} pour TOUS les jours de la fenetre, meme les jours vides.

    Un jour sans drop reste dans le dictionnaire avec une liste vide : c'est la
    grille qui doit etre complete, pas seulement les cases pleines.
    """
    calendrier = {j: Jour(j) for j in jours_de(debut, fin)}
    par_jour: Dict[_dt.date, Dict[str, Serie]] = {j: {} for j in calendrier}

    for ligne in lignes:
        jour = date_de(ligne.get("releaseDate"))
        if jour is None or jour < debut or jour > fin:
            continue
        cle = cle_serie(ligne)
        if not cle:
            continue
        rang = ORDRE_RARETE.get(str(ligne.get("rarity") or "").strip().upper(), 999)
        image = str(ligne.get("image_url") or "").strip()
        seau = par_jour[jour]
        serie = seau.get(cle)
        if serie is None:
            serie = Serie(
                cle=cle,
                nom=str(ligne.get("veve_series_name") or ligne.get("name") or "").strip(),
                marque=str(ligne.get("veve_brand") or "").strip(),
                licensor=str(ligne.get("veve_licensor") or "").strip(),
                nature=ligne.get("_nature", ""),
                image_url=image,
                drop_method=str(ligne.get("drop_method") or "").strip(),
                rang_image=rang if image else 999,
            )
            seau[cle] = serie
        elif image and rang < serie.rang_image:
            # L'image de la rarete la plus BASSE gagne : couverture de base.
            serie.image_url, serie.rang_image = image, rang
        serie.editions += 1

    for jour, seau in par_jour.items():
        # Tri stable et lisible : les series illustrees d'abord (une vignette
        # sans image est un trou dans la grille), puis par nom.
        calendrier[jour].series = sorted(
            seau.values(), key=lambda s: (0 if s.image_url else 1, s.nom.lower()))
    return calendrier


def charger(debut: _dt.date, fin: _dt.date, *, xlsx: Optional[str] = None,
            sheet: bool = False) -> Dict[_dt.date, Jour]:
    return grouper(lignes_sheet() if sheet else lignes_xlsx(xlsx or ""), debut, fin)


# ------------------------------------------------------------------- affichages

MOIS_FR_COURT = ("JANV.", "FÉVR.", "MARS", "AVR.", "MAI", "JUIN", "JUIL.",
                 "AOÛT", "SEPT.", "OCT.", "NOV.", "DÉC.")
JOURS_FR = ("LUN", "MAR", "MER", "JEU", "VEN", "SAM", "DIM")


def libelle_periode(debut: _dt.date, fin: _dt.date, mois=None, lien: str = " » ") -> str:
    """« 29 JUIN » 2 AOÛT 2026 » — sans jamais dependre de la locale du systeme.

    ⚠️ Deux pieges cousus ici :
      * `strftime('%B')` rendrait « June » sur une machine en anglais : les noms
        de mois sont ecrits en dur ;
      * la LANGUE suit le THEME, pas la machine — VeVe Insights ecrit en anglais.
        D'ou `mois`, la table de mois du theme (defaut : francais).
    """
    table = mois or MOIS_FR_COURT
    return "%d %s%s%d %s %d" % (debut.day, table[debut.month - 1], lien,
                                fin.day, table[fin.month - 1], fin.year)
